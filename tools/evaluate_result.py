# -*- coding: utf-8 -*-

# @Time:      2022/11/1 15:40
# @Author:    qiao
# @Email:     rukunqiao@outlook.com
# @File:      evaluate_result.py
# @Software:  PyCharm
# @Description:
#   Automatically output the results into an Excel file.
#   Notice the coordinate for open3d is: left x, up y, left handed.
#   TODO: Provide an example file and script.

# - Package Imports - #
import time
from configparser import ConfigParser
from pathlib import Path
import numpy as np
import cv2
import torch
import openpyxl
import open3d as o3d
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Alignment
import shutil

import pointerlib as plb


# - Coding Part - #
class Evaluator:
    def __init__(self, workbook, flush_flag, pre_process_flag=False):
        self.workbook = openpyxl.load_workbook(str(workbook))
        self.workbook_path = workbook
        self.flush_flag = flush_flag
        self.pre_process_flag = pre_process_flag
        self.calib_para = None
        self.visualizer = None

    def convert_to_depth(self, out_path):
        for ply_file in tqdm(list(out_path.rglob('*.ply')), desc='mesh2depth'):
            depth_file = ply_file.parent / 'depth.png'
            if depth_file.exists() and not self.flush_flag:
                continue
            else:
                self.visualizer.set_mesh(ply_file)
                self.visualizer.get_depth(depth_file)
        pass

    def flush_depth_est_shape(self, cmp_set):
        depth_gt_file, depth_res_file, mask_gt_file = cmp_set
        depth_gt = plb.imload(depth_gt_file, scale=10.0)
        depth_res = plb.imload(depth_res_file, scale=10.0)

        if depth_res.shape[-2] != depth_gt.shape[-2] or depth_res.shape[-1] != depth_gt.shape[-1]:
            mask = plb.imload(mask_gt_file)
            h_src, w_src = map(lambda x: x.min().item(), torch.where(mask[0] > 0))
            h_end, w_end = map(lambda x: x.max().item(), torch.where(mask[0] > 0))
            depth_res_re = torch.zeros_like(depth_gt)
            depth_res_re[:, h_src:h_end, w_src:w_end] = depth_res
            depth_res = depth_res_re
            plb.imsave(depth_res_file, depth_res, scale=10.0, img_type=np.uint16)

        pass

    def evaluate_sequence(self, scene_name):
        work_sheet = self.workbook[scene_name]
        data_path = Path(work_sheet['B1'].value)
        out_path = Path(work_sheet['B2'].value)

        # load parameters
        config = ConfigParser()
        config.read(str(data_path / 'config.ini'), encoding='utf-8')
        self.calib_para = config['RawCalib']
        self.visualizer = plb.DepthVisualizer(
            img_size=plb.str2tuple(self.calib_para['img_size'], item_type=int),
            img_intrin=plb.str2tuple(self.calib_para['img_intrin'], item_type=float),
            pos=[-50.0, 50.0, 50.0],
        )
        if self.pre_process_flag:
            self.convert_to_depth(out_path)

        # Get all exp_tags
        start_row = 5
        cmp_sets = []

        current_row = start_row
        exp_set = sorted([x for x in out_path.glob('*') if x.is_dir()])
        for i, exp_path in enumerate(exp_set):
            epoch_set = sorted([x for x in (exp_path / 'output').glob('*_*') if x.is_dir()],
                               key=lambda x: int(x.name.split('_')[1]))
            for j, epoch_folder in enumerate(epoch_set):
                work_sheet.cell(current_row, 1).value = exp_path.name
                work_sheet.cell(current_row, 2).value = str(epoch_folder)
                work_sheet.cell(current_row, 3).value = int(epoch_folder.name.split('_')[1])
                cmp_set = self._build_up_cmp_set(data_path, epoch_folder)
                if self.pre_process_flag:
                    self.flush_depth_est_shape(cmp_set)
                cmp_sets.append(cmp_set)
                current_row += 1
        total_row = current_row
        # if self.pre_process_flag:
        #     map(self.flush_depth_est_shape, cmp_sets)

        # Evaluate: column = 4,5,6,7
        for row_idx in tqdm(range(start_row, total_row)):
            depth_gt, depth_est, mask_gt = cmp_sets[row_idx - start_row]
            exp_tag = work_sheet.cell(row_idx, 1).value
            epoch_dir = Path(work_sheet.cell(row_idx, 2).value)
            self.visualizer.set_depth(depth_est, mask_gt)
            if self.visualizer.depth.max() < 5.0:
                self.visualizer.depth[:, :] = 700.0

            # Check if depth2pcd or vis is needed.
            if self.flush_flag or not (epoch_dir / 'pcd.asc').exists() or not (epoch_dir / 'vis.png').exists():
                self.visualizer.update()

                if self.flush_flag or not (epoch_dir / 'pcd.asc').exists():
                    self.visualizer.get_pcd(epoch_dir / 'pcd.asc')

                # Check if depth vis is needed.
                if self.flush_flag or not (epoch_dir / 'vis.png').exists():
                    self.visualizer.get_view(epoch_dir / 'vis.png')

            # Evaluate here.
            res, diff, mask = self._evaluate_exp_outs(cmp_sets[row_idx - start_row])

            # Check if step vis is needed.
            if self.flush_flag or not (epoch_dir / 'step_vis.png').exists():
                step_vis = self._draw_step_vis(diff, mask)
                plb.imsave(epoch_dir / 'step_vis.png', step_vis)

            # Write
            work_sheet.cell(row_idx, 4).value = float(f'{res[0] * 100.0:.2f}')
            work_sheet.cell(row_idx, 5).value = float(f'{res[1] * 100.0:.2f}')
            work_sheet.cell(row_idx, 6).value = float(f'{res[2] * 100.0:.2f}')
            work_sheet.cell(row_idx, 7).value = float(f'{res[3]:.3f}')

        self.workbook.save(self.workbook_path)
        pass

    def _build_up_cmp_set(self, data_path, out_path):
        # cmp_set = []

        depth_gt_file = data_path / 'gt' / 'depth.png'
        mask_gt_file = data_path / 'gt' / 'mask_occ.png'
        depth_res_file = out_path / 'depth.png'

        return depth_gt_file, depth_res_file, mask_gt_file

    def _evaluate_exp_outs(self, cmp_set):

        depth_gt, depth_res, mask_gt = cmp_set
        depth_gt = plb.imload(depth_gt, scale=10.0)
        depth_res = plb.imload(depth_res, scale=10.0)
        # mask_gt = plb.imload(mask_gt)
        mask_gt = (depth_gt > 0.0).float()

        diff = (depth_gt - depth_res)
        diff_vec = diff[mask_gt > 0.0]
        total_num = diff_vec.shape[0]
        err10_num = (torch.abs(diff_vec) > 1.0).float().sum() / total_num
        err20_num = (torch.abs(diff_vec) > 2.0).float().sum() / total_num
        err50_num = (torch.abs(diff_vec) > 5.0).float().sum() / total_num
        avg = torch.abs(diff_vec).sum() / total_num
        res_array = np.array([err10_num, err20_num, err50_num, avg])

        return res_array, diff, mask_gt

    def _ply2depth(self, ply_file, visible=False):
        wid, hei = plb.str2tuple(self.calib_para['img_size'], item_type=int)
        fx, fy, dx, dy = plb.str2tuple(self.calib_para['img_intrin'], item_type=float)
        vis = o3d.visualization.Visualizer()
        vis.create_window(width=wid, height=hei, visible=visible)
        inf_pinhole = o3d.camera.PinholeCameraIntrinsic(
            width=wid, height=hei,
            fx=fx, fy=fy, cx=dx, cy=dy
        )
        camera = o3d.camera.PinholeCameraParameters()
        camera.intrinsic = inf_pinhole
        camera.extrinsic = np.eye(4, dtype=np.float64)
        ctr = vis.get_view_control()
        ctr.convert_from_pinhole_camera_parameters(camera, allow_arbitrary=True)

        vis.poll_events()
        vis.update_renderer()
        time.sleep(0.01)

        mesh = o3d.io.read_triangle_mesh(str(ply_file))

        # Coordinate change
        vertices_np = np.asarray(mesh.vertices)
        mesh.vertices = o3d.utility.Vector3dVector(vertices_np)

        mesh.compute_vertex_normals()

        vis.add_geometry(mesh)
        ctr.convert_from_pinhole_camera_parameters(camera, allow_arbitrary=True)

        vis.poll_events()
        vis.update_renderer()
        time.sleep(0.01)

        # vis.run()

        depth = vis.capture_depth_float_buffer()
        depth = np.asarray(depth).copy()
        depth[np.isnan(depth)] = 0.0

        img = vis.capture_screen_float_buffer()
        img = np.asarray(img).copy()

        vis.clear_geometries()
        vis.destroy_window()

        return depth, img

    def _depth2pcd(self, depth_file, mask_file):
        img_intrin = plb.str2tuple(self.calib_para['img_intrin'], item_type=float)
        img_size = plb.str2tuple(self.calib_para['img_size'], int)
        depth_viz = plb.DepthVisualizer(img_size, img_intrin)
        depth_viz.set_depth(depth_file, mask_file)
        depth_viz.update()
        xyz_set = depth_viz.get_pcd()
        xyz_set = xyz_set[xyz_set[:, -1] > 10.0, :]
        return xyz_set

    def _pcd2vis(self, pcd_file, visible=False):
        wid, hei = plb.str2tuple(self.calib_para['img_size'], item_type=int)
        fx, fy, dx, dy = plb.str2tuple(self.calib_para['img_intrin'], item_type=float)
        vis = o3d.visualization.Visualizer()
        vis.create_window(width=wid, height=hei, visible=visible)
        inf_pinhole = o3d.camera.PinholeCameraIntrinsic(
            width=wid, height=hei,
            fx=fx, fy=fy, cx=dx, cy=dy
        )
        camera = o3d.camera.PinholeCameraParameters()
        camera.intrinsic = inf_pinhole
        camera.extrinsic = np.eye(4, dtype=np.float64)
        ctr = vis.get_view_control()
        ctr.convert_from_pinhole_camera_parameters(camera, allow_arbitrary=True)

        vis.poll_events()
        vis.update_renderer()
        time.sleep(0.01)

        # Create pcd
        xyz_set = np.loadtxt(str(pcd_file), delimiter=',', encoding='utf-8')
        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(xyz_set)
        pcd.estimate_normals(
            search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=30.0, max_nn=30)
        )
        pcd.colors = o3d.utility.Vector3dVector(np.ones_like(xyz_set) * 0.5)

        vis.add_geometry(pcd)
        ctr.convert_from_pinhole_camera_parameters(camera, allow_arbitrary=True)

        vis.poll_events()
        vis.update_renderer()
        time.sleep(0.01)

        # vis.run()

        img = vis.capture_screen_float_buffer()
        img = np.asarray(img).copy()

        vis.clear_geometries()
        vis.destroy_window()

        return img

    def _draw_step_vis(self, diff, mask):
        step_err = torch.ones_like(diff)
        step_err[torch.abs(diff) > 1.0] = 2.0
        step_err[torch.abs(diff) > 2.0] = 3.0
        step_err[torch.abs(diff) > 5.0] = 4.0
        step_vis = plb.VisualFactory.err_visual(step_err, mask, max_val=4.0, color_map=cv2.COLORMAP_WINTER)
        step_vis = cv2.cvtColor(plb.t2a(step_vis), cv2.COLOR_BGR2RGB)
        return step_vis

    def sum_average(self, scene_name):
        src_work_sheet = self.workbook[scene_name]
        dst_work_sheet = self.workbook[f'{scene_name}-Sum']

        # Get all exp_tags & values
        start_row = 5
        raw_results = []
        for row_idx in range(start_row, src_work_sheet.max_row + 1):
            raw_results.append([
                src_work_sheet.cell(row_idx, 1).value,
                [
                    float(src_work_sheet.cell(row_idx, 3).value),
                    float(src_work_sheet.cell(row_idx, 4).value),
                    float(src_work_sheet.cell(row_idx, 5).value),
                    float(src_work_sheet.cell(row_idx, 6).value),
                ]
            ])

        # Group
        grouped_results = {}
        for exp_tag, eval_res in raw_results:
            exp_key = exp_tag
            if '_exp' in exp_tag:
                exp_key = exp_tag.split('_exp')[0]
            if exp_key not in grouped_results:
                grouped_results[exp_key] = []
            grouped_results[exp_key].append(np.array(eval_res))

        # Put into new sheet
        for i, exp_tag in enumerate(grouped_results):
            dst_work_sheet.cell(2 + i, 1).value = exp_tag
            exp_res = np.stack(grouped_results[exp_tag], axis=0)
            exp_avg = np.average(exp_res, axis=0)
            dst_work_sheet.cell(2 + i, 2).value = f'{exp_avg[0]:.2f}'
            dst_work_sheet.cell(2 + i, 3).value = f'{exp_avg[1]:.2f}'
            dst_work_sheet.cell(2 + i, 4).value = f'{exp_avg[2]:.2f}'
            dst_work_sheet.cell(2 + i, 5).value = f'{exp_avg[3]:.3f}'

        self.workbook.save(self.workbook_path)
        pass


def mesh_color_visualize(depth, mask, config, img_list, out_folder):
    """Given the depth and projected image, generate view with color."""

    def safe_load(file_path, *args, **kwargs):
        if isinstance(file_path, Path) or isinstance(file_path, str):
            return plb.imload(file_path, *args, **kwargs)
        else:
            return file_path

    depth = safe_load(depth, 10.0, flag_tensor=False)
    mask = safe_load(mask, flag_tensor=False)
    for i in range(len(img_list)):
        img_list[i] = safe_load(img_list[i], flag_tensor=False)
    cfg = ConfigParser()
    cfg.read(str(config), encoding='utf-8')
    img_size = plb.str2tuple(cfg['RawCalib']['img_size'], item_type=int)
    img_intrin = plb.str2tuple(cfg['RawCalib']['img_intrin'], item_type=float)
    pos = [-50.0, 50.0, 50.0]

    # depth -> pcd & color
    wid, hei = img_size
    fx, fy, dx, dy = img_intrin
    hh = np.arange(0, hei).reshape(-1, 1).repeat(wid, axis=1)
    ww = np.arange(0, wid).reshape(1, -1).repeat(hei, axis=0)
    xyz_mat = np.stack([
        (ww - dx) / fx,
        (hh - dy) / fy,
        np.ones_like(ww)
    ], axis=0) * depth[None]  # [3, H, W]
    xyz_set = xyz_mat.reshape(3, -1).transpose()    # [H * W, 3]
    pcd = xyz_set[mask.reshape(-1) > 0.0, :]        # [N, 3]

    color_set = []
    for img in img_list:
        if len(img.shape) == 2:
            img = np.stack([img] * 3, axis=0)
        img_vec = img.reshape(img.shape[0], -1).transpose()
        color_set.append(img_vec[mask.reshape(-1) > 0.0, :])  # [N, 3]

    view_set = []
    with plb.MyO3DVisualizer(img_size, img_intrin, pos, True) as v:
        for color_vec in color_set:
            pcd_o3d = o3d.geometry.PointCloud()
            pcd_o3d.points = o3d.utility.Vector3dVector(pcd)
            pcd_o3d.estimate_normals(
                search_param=o3d.geometry.KDTreeSearchParamHybrid(radius=30.0, max_nn=30)
            )
            pcd_o3d.colors = o3d.utility.Vector3dVector(color_vec)
            v.vis.add_geometry(pcd_o3d)
            v.flush_pos()
            v.update()
            view_set.append(v.get_screen())
            v.vis.remove_geometry(pcd_o3d)
        time.sleep(1)

    out_folder.mkdir(exist_ok=True, parents=True)
    for i, view in enumerate(view_set):
        plb.imsave(out_folder / f'vis_{i}.png', view)

    pass


def export_text_to_file(workbook, params):
    sheet_names = ['scene_00', 'scene_01', 'scene_02', 'scene_03']
    pat_set = [7, 6, 5, 4, 3]
    row_pair = [
        ('7', 6),
        ('6', 7),
        ('5', 8),
        ('4', 9),
        ('3', 10),
        ('23456', 11),
        ('12457', 12),
        ('1346', 16),
        ('2457', 15),
    ]
    res = []
    for pat_num, row_idx in row_pair:
        for i, exp_name in enumerate(['Naive', 'NeRF', 'NeuS', 'Ours']):
            title = f'{pat_num}-{exp_name}: & '
            nums = []
            col_s = i * 4 + 2
            for x, sheet_name in enumerate(sheet_names):
                for bias in [2, 3]:
                    val = workbook[sheet_name].cell(row=row_idx, column=col_s + bias).value
                    if val is None:
                        nums.append('~')
                    else:
                        if exp_name == 'Ours':
                            val = '\\textbf{' + val + '}'
                        nums.append(str(val))
            num_str = ' & '.join(nums)
            res.append(title + num_str + '\n')
    with open(params['txt_name'], 'w+') as file:
        file.writelines(res)

    print(len(res))
    pass


def main():

    app = Evaluator(
        workbook='C:/SLDataSet/SLNeRF/result_3DV.xlsx',
        flush_flag=False,
        pre_process_flag=True,
    )
    app.evaluate_sequence('scene_00')
    app.evaluate_sequence('scene_01')
    app.evaluate_sequence('scene_02')
    app.evaluate_sequence('scene_03')

    # data_path = Path('C:/SLDataSet/SLNeRF/7_Dataset0531/scene_00')
    # out_path = Path('C:/SLDataSet/SLNeRF/7_Dataset0531-out/scene_00')
    # res_path = out_path / 'xyz2sdf-gc012345wrp' / 'output' / 'e_50000'
    # mesh_color_visualize(
    #     depth=res_path / 'depth.png',
    #     mask=data_path / 'gt' / 'mask_occ.png',
    #     config=data_path / 'config.ini',
    #     img_list=[data_path / 'img' / f'img_gc{x}.png' for x in ['2', '3', '4', '5']],
    #     out_folder=res_path / 'view_vis',
    # )

    # app.sum_average('NonRigidReal')


def main_sup():
    main_folder = Path(r'C:\Users\qiao\Desktop\CVPR2023_Sub')
    res_folder = main_folder / 'gif_set'
    res_folder.mkdir(parents=True, exist_ok=True)
    pcd_list = [
        # (main_folder / f'scene_00' / 'gt.asc', res_folder / 'gt_00')
    ]
    for scene_idx in [0]:
        for exp_set in ['NeuS']:  # [NeuS] 'Ours', 'GrayCode', 'Ours-sample', 'Ours-warp'
            for exp_tag in ['pat5']:
                pcd_list.append(
                    (main_folder / f'scene_{scene_idx:02}' / exp_set / f'{exp_tag}.ply',
                     res_folder / f'{exp_set}_{scene_idx:02}_{exp_tag}')
                )
    json_file = Path(r'C:\SLDataSet\20220907real\open3dvis.json')

    wid, hei = 640, 480
    total_num = 128
    vis = o3d.visualization.Visualizer()
    vis.create_window(width=wid, height=hei)
    vis.get_render_option().point_size = 1.0
    ctr = vis.get_view_control()
    param = o3d.io.read_pinhole_camera_parameters(str(json_file))

    # Create pos set
    pos_list = []
    current_pos = param.extrinsic[:3, -1]
    # xyz_set = np.loadtxt(pcd_list[0][0], delimiter=',', encoding='utf-8')
    # mid_point = (xyz_set.max(axis=0) + xyz_set.min(axis=0)) * 0.5
    # mid_point = - mid_point
    # diff_vec = current_pos - mid_point
    # diff_vec[1] = 0.0
    # rad = np.linalg.norm(diff_vec)
    rad = 100
    for frm_idx in range(total_num):
        angle = np.pi * 2 / (total_num * 0.5) * frm_idx
        # pos = mid_point + np.array([
        #     np.cos(angle) * rad,
        #     0,
        #     np.sin(angle) * rad
        # ], dtype=mid_point.dtype)
        pos = current_pos + np.array([
            np.cos(angle) * rad,
            np.sin(angle) * rad,
            np.cos(angle * 1.5) * rad,
        ])
        cam_pos = plb.PosManager()
        # cam_pos.set_from_look(pos, look, up)
        cam_pos.set_rot(mat=param.extrinsic[:3, :3])
        cam_pos.set_trans(pos)
        pos_list.append(cam_pos)

    for pcd_path, save_folder in pcd_list:
        if not pcd_path.exists():
            continue
        # try:
        #     xyz_set = np.loadtxt(str(pcd_path), delimiter=',', encoding='utf-8')
        # except ValueError as e:
        #     xyz_set = np.loadtxt(str(pcd_path), encoding='utf-8')
        # xyz_set = xyz_set[xyz_set.sum(axis=1) > 1.0]
        # pcd = o3d.geometry.PointCloud()
        # pcd.points = o3d.utility.Vector3dVector(xyz_set)
        # pcd.estimate_normals()
        # pcd.colors = o3d.utility.Vector3dVector(np.ones_like(np.asarray(pcd.points)) * 0.5)
        pcd = o3d.io.read_triangle_mesh(str(pcd_path))
        pcd.compute_vertex_normals()
        pcd.compute_triangle_normals()
        vis.add_geometry(pcd)
        save_folder.mkdir(exist_ok=True, parents=True)
        for frm_idx in range(total_num):
            # Ctr
            param.extrinsic = pos_list[frm_idx].get_4x4mat()
            ctr.convert_from_pinhole_camera_parameters(param)
            vis.poll_events()
            vis.update_renderer()
            time.sleep(0.01)
            image = vis.capture_screen_float_buffer()
            image = np.asarray(image).copy()
            plb.imsave(save_folder / f'{frm_idx}.png', image)
        vis.clear_geometries()

    pass


def draw_gifs():
    main_folder = Path(r'C:\Users\qiao\Desktop\CVPR2023_Sub\Supplementary\data_scene0')
    sub_folders = [x for x in main_folder.glob('*') if x.is_dir()]
    for sub_folder in sub_folders:
        gif_name = sub_folder.parent / f'{sub_folder.name}.gif'
        writer = plb.GifWriter(gif_name, fps=2)
        writer.run_folder(sub_folder)


def clean_asc():
    main_folder = Path(r'C:\Users\qiao\Desktop\CVPR2023_Sub\results')
    asc_files = [x for x in main_folder.glob('*.asc')]
    for asc_file in asc_files:
        try:
            xyz_set = np.loadtxt(str(asc_file), delimiter=',', encoding='utf-8')
        except ValueError as e:
            xyz_set = np.loadtxt(str(asc_file), encoding='utf-8')
        xyz_set = xyz_set[xyz_set.sum(axis=1) > 1.0]
        np.savetxt(str(asc_file.parent / f'{asc_file.stem}.xyz'), xyz_set,
                   fmt='%.2f', delimiter=' ', newline='\n', encoding='utf-8')


if __name__ == '__main__':
    # test()
    main()
    # main_sup()
    # draw_gifs()
